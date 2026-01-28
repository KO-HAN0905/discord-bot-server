# -*- coding: utf-8 -*-
import discord
from discord.ext import commands, tasks
from discord.ui import Modal, TextInput, View, Button
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import json
import os
import datetime as dt
from datetime import datetime, timedelta
from contextlib import suppress

# D-Day 추가 모달
class DDayModal(Modal, title="D-Day 추가"):
    """D-Day 추가 모달 UI"""
    
    name = TextInput(
        label="D-Day 이름",
        placeholder="예: 시험, 생일, 프로젝트마감일",
        required=True,
        min_length=1,
        max_length=50
    )
    
    date = TextInput(
        label="목표 날짜",
        placeholder="YYYY-MM-DD (예: 2026-03-15)",
        required=True,
        min_length=10,
        max_length=10
    )
    
    message = TextInput(
        label="메시지 (선택사항)",
        placeholder="이 D-Day에 대한 설명을 입력하세요",
        required=False,
        style=discord.TextStyle.paragraph,
        max_length=200
    )
    
    def __init__(self, cog):
        super().__init__()
        self.cog = cog
    
    async def on_submit(self, interaction: discord.Interaction):
        """모달 제출 처리"""
        try:
            # 날짜 형식 검증
            target_date = datetime.strptime(self.date.value, "%Y-%m-%d")
            
            # 엑셀에 데이터 추가
            wb = openpyxl.load_workbook(self.cog.excel_file)
            ws = wb.active
            
            row = ws.max_row + 1
            ws[f"A{row}"] = self.name.value
            ws[f"B{row}"] = target_date.date()
            ws[f"C{row}"] = self.cog.calculate_dday(target_date)
            ws[f"D{row}"] = datetime.now().date()
            ws[f"E{row}"] = f"활성 - {self.message.value}" if self.message.value else "활성"
            
            # 셀 스타일
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            wb.save(self.cog.excel_file)
            
            dday_count = self.cog.calculate_dday(target_date)
            
            embed = discord.Embed(
                title="✅ D-Day 추가됨",
                color=discord.Color.green()
            )
            embed.add_field(name="이름", value=self.name.value, inline=True)
            embed.add_field(name="날짜", value=self.date.value, inline=True)
            embed.add_field(
                name="D-Day",
                value=f"D{dday_count}" if dday_count < 0 else f"D+{dday_count}",
                inline=True
            )
            if self.message.value:
                embed.add_field(name="메시지", value=self.message.value, inline=False)
            
            await interaction.response.send_message(embed=embed, ephemeral=True)
            
            # 채널 업데이트
            await self.cog.refresh_all_dday_channels()
        except ValueError:
            embed = discord.Embed(
                title="❌ 오류",
                description="날짜 형식이 잘못되었습니다.\n올바른 형식: YYYY-MM-DD (예: 2026-03-15)",
                color=discord.Color.red()
            )
            await interaction.response.send_message(embed=embed, ephemeral=True)
        except Exception as e:
            embed = discord.Embed(
                title="❌ 오류",
                description=f"오류가 발생했습니다: {str(e)}",
                color=discord.Color.red()
            )
            await interaction.response.send_message(embed=embed, ephemeral=True)

# D-Day UI 버튼 뷰
class DDayView(View):
    """D-Day 관리 UI 버튼"""
    
    def __init__(self, cog):
        super().__init__(timeout=300)  # 5분 타임아웃
        self.cog = cog
    
    @discord.ui.button(label="➕ D-Day 추가", style=discord.ButtonStyle.success, row=0)
    async def add_dday_button(self, interaction: discord.Interaction, button: Button):
        """D-Day 추가 버튼"""
        modal = DDayModal(self.cog)
        await interaction.response.send_modal(modal)
    
    @discord.ui.button(label="📋 목록 보기", style=discord.ButtonStyle.primary, row=0)
    async def list_button(self, interaction: discord.Interaction, button: Button):
        """목록 보기 버튼"""
        try:
            wb = openpyxl.load_workbook(self.cog.excel_file)
            ws = wb.active
            
            if ws.max_row <= 1:
                await interaction.response.send_message("등록된 D-Day가 없습니다.", ephemeral=True)
                return
            
            embed = discord.Embed(title="📅 D-Day 목록", color=discord.Color.blue())
            
            for row in range(2, ws.max_row + 1):
                if not (name := ws[f"A{row}"].value):
                    continue

                date = ws[f"B{row}"].value or ""
                dday = ws[f"C{row}"].value
                status = ws[f"E{row}"].value or ""

                if dday < 0:
                    text = f"📍 {dday}일 (경과)"
                elif dday == 0:
                    text = "🎉 D-Day!"
                else:
                    text = f"⏳ D+{dday}일"

                message = status.split(" - ", 1)[1] if " - " in status else ""
                value = f"{text}\n목표: {date}"
                if message:
                    value += f"\n💬 {message}"

                embed.add_field(name=name, value=value, inline=False)
            
            await interaction.response.send_message(embed=embed, ephemeral=True)
        except Exception as e:
            await interaction.response.send_message(f"❌ 오류: {e}", ephemeral=True)
    
    @discord.ui.button(label="🔄 새로고침", style=discord.ButtonStyle.secondary, row=0)
    async def refresh_button(self, interaction: discord.Interaction, button: Button):
        """새로고침 버튼"""
        try:
            wb = openpyxl.load_workbook(self.cog.excel_file)
            ws = wb.active
            
            for row in range(2, ws.max_row + 1):
                if date_cell := ws[f"B{row}"].value:
                    if isinstance(date_cell, str):
                        date_cell = datetime.strptime(date_cell, "%Y-%m-%d")
                    ws[f"C{row}"] = self.cog.calculate_dday(date_cell)
            
            wb.save(self.cog.excel_file)
            
            embed = discord.Embed(
                title="✅ D-Day가 새로고침되었습니다",
                color=discord.Color.green()
            )
            await interaction.response.send_message(embed=embed, ephemeral=True)
        except Exception as e:
            await interaction.response.send_message(f"❌ 오류: {e}", ephemeral=True)
    
    @discord.ui.button(label="❌ 삭제", style=discord.ButtonStyle.red, row=0)
    async def delete_button(self, interaction: discord.Interaction, button: Button):
        """삭제 선택 버튼 - 별도 처리 필요"""
        await interaction.response.send_message(
            "삭제할 D-Day 이름을 입력하세요:\n`!디데이 삭제 <이름>`",
            ephemeral=True
        )

    @discord.ui.button(label="📂 채널 생성/동기화", style=discord.ButtonStyle.secondary, row=1)
    async def sync_channels_button(self, interaction: discord.Interaction, button: Button):
        """D-Day 채널 생성 및 동기화"""
        if not interaction.guild:
            await interaction.response.send_message("❌ 서버에서만 사용할 수 있습니다.", ephemeral=True)
            return
        if not interaction.user.guild_permissions.manage_channels:
            await interaction.response.send_message("❌ 채널 생성/동기화는 관리자만 사용할 수 있습니다.", ephemeral=True)
            return
        await interaction.response.defer(ephemeral=True, thinking=True)
        await self.cog.create_dday_channels(interaction.guild)
        await interaction.followup.send("✅ D-Day 채널이 생성/동기화되었습니다.", ephemeral=True)

    @discord.ui.button(label="🔁 채널 새로고침", style=discord.ButtonStyle.secondary, row=1)
    async def refresh_channels_button(self, interaction: discord.Interaction, button: Button):
        """D-Day 채널 이름 새로고침"""
        if not interaction.guild:
            await interaction.response.send_message("❌ 서버에서만 사용할 수 있습니다.", ephemeral=True)
            return
        if not interaction.user.guild_permissions.manage_channels:
            await interaction.response.send_message("❌ 채널 새로고침은 관리자만 사용할 수 있습니다.", ephemeral=True)
            return
        await interaction.response.defer(ephemeral=True, thinking=True)
        await self.cog.refresh_all_dday_channels()
        await interaction.followup.send("✅ D-Day 채널이 새로고침되었습니다.", ephemeral=True)

class DDayManager(commands.Cog):
    def __init__(self, bot):
        self.bot = bot
        self.excel_file = "data/dday.xlsx"
        self.settings_file = "data/settings.json"
        self.dday_channels = {}  # 각 D-day별 채널 저장
        self.dday_category = None  # D-Day 카테고리
        self.ensure_excel()
        self.update_dday_channels.start()  # 매일 자정 업데이트 시작

    def cog_unload(self):
        """Cog 언로드 시 작업 중지"""
        self.update_dday_channels.cancel()

    @tasks.loop(time=dt.time(hour=0, minute=0))
    async def update_dday_channels(self):
        """매일 자정에 D-Day 채널 업데이트"""
        await self.refresh_all_dday_channels()
    
    @update_dday_channels.before_loop
    async def before_update(self):
        """봇이 준비될 때까지 대기"""
        await self.bot.wait_until_ready()
        # 첫 실행 시 즉시 업데이트
        await self.refresh_all_dday_channels()
    
    async def create_dday_channels(self, guild):
        """D-Day 카테고리 및 채널 생성"""
        # D-Day 카테고리 찾기 또는 생성
        self.dday_category = discord.utils.get(guild.categories, name="📅 D-DAY")
        if not self.dday_category:
            self.dday_category = await guild.create_category("📅 D-DAY")
        
        # D-Day 목록 가져오기
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            name = ws[f"A{row}"].value
            if name and name not in self.dday_channels:
                # 음성 채널 생성 (이름 표시용)
                channel = await guild.create_voice_channel(
                    name="🔄 계산중...",
                    category=self.dday_category,
                    user_limit=0  # 입장 불가
                )
                self.dday_channels[name] = channel
        
        # 즉시 업데이트
        await self.refresh_all_dday_channels()
    
    async def refresh_all_dday_channels(self):
        """모든 D-Day 채널 업데이트"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active

            entries = []  # (name, dday_value)

            # D-Day 값 계산 및 저장
            for row in range(2, ws.max_row + 1):
                if not (name := ws[f"A{row}"].value):
                    continue

                dday_value = None
                if date_cell := ws[f"B{row}"].value:
                    if isinstance(date_cell, str):
                        date_cell = datetime.strptime(date_cell, "%Y-%m-%d")
                    dday_value = self.calculate_dday(date_cell)
                    ws[f"C{row}"] = dday_value

                entries.append((name, dday_value))

            wb.save(self.excel_file)

            # 길드(서버) 정보 가져오기
            guild = None
            if self.dday_category and self.dday_category.guild:
                guild = self.dday_category.guild
            elif self.dday_channels:
                first_channel = next(iter(self.dday_channels.values()), None)
                if first_channel:
                    guild = first_channel.guild
            
            # 길드가 없으면 봇이 속한 첫 번째 길드 사용
            if not guild and self.bot.guilds:
                guild = self.bot.guilds[0]
                # 카테고리 찾기
                self.dday_category = discord.utils.get(guild.categories, name="📅 D-DAY")

            for name, dday_value in entries:
                # 채널 확보
                channel = self.dday_channels.get(name)
                if not channel and guild and self.dday_category:
                    with suppress(Exception):
                        channel = await guild.create_voice_channel(
                            name="🔄 계산중...",
                            category=self.dday_category,
                            user_limit=0
                        )
                        self.dday_channels[name] = channel

                if channel and dday_value is not None:
                    if dday_value < 0:
                        channel_name = f"✅ {name}: D{dday_value}"
                    elif dday_value == 0:
                        channel_name = f"🎉 {name}: D-DAY!"
                    else:
                        channel_name = f"📅 {name}: D-{dday_value}"

                    with suppress(Exception):
                        await channel.edit(name=channel_name)
                        print(f"D-Day 채널 업데이트 성공: {channel_name}")

        except Exception as e:
            print(f"D-Day 채널 업데이트 오류: {e}")
            import traceback
            traceback.print_exc()

    def ensure_excel(self):
        """엑셀 파일 초기화"""
        os.makedirs(os.path.dirname(self.excel_file), exist_ok=True)
        
        if not os.path.exists(self.excel_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "D-Day"
            
            # 헤더 설정
            headers = ["이름", "목표날짜", "D-Day", "생성일", "상태"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            wb.save(self.excel_file)

    @commands.group(name="디데이", help="D-Day 관리")
    async def dday(self, ctx):
        if ctx.invoked_subcommand is None:
            embed = discord.Embed(title="📅 D-Day 도움말", color=discord.Color.blue())
            embed.add_field(name="UI 관리", value="`!디데이 ui` - 버튼식 UI 사용", inline=False)
            embed.add_field(name="채널 생성", value="`!디데이 채널생성` - D-Day 채널 카테고리 생성", inline=False)
            embed.add_field(name="채널 업데이트", value="`!디데이 채널업데이트` - D-Day 채널 정보 갱신", inline=False)
            embed.add_field(name="추가", value="`!디데이 추가 <이름> <날짜(YYYY-MM-DD)>`", inline=False)
            embed.add_field(name="삭제", value="`!디데이 삭제 <이름>`", inline=False)
            embed.add_field(name="목록", value="`!디데이 목록`", inline=False)
            embed.add_field(name="공지", value="`!디데이 공지 <채널>`", inline=False)
            await ctx.send(embed=embed)
    
    @dday.command(name="채널생성", help="D-Day 채널 생성")
    @commands.has_permissions(administrator=True)
    async def create_channels(self, ctx):
        """D-Day 카테고리 및 채널 생성"""
        await ctx.send("📅 D-Day 채널을 생성하는 중...")
        await self.create_dday_channels(ctx.guild)
        await ctx.send("✅ D-Day 채널이 생성되었습니다!")
    
    @dday.command(name="채널업데이트", help="D-Day 채널 업데이트")
    @commands.has_permissions(administrator=True)
    async def update_channels(self, ctx):
        """D-Day 채널 정보 갱신"""
        await ctx.send("🔄 D-Day 채널을 업데이트하는 중...")
        await self.refresh_all_dday_channels()
        await ctx.send("✅ D-Day 채널이 업데이트되었습니다!")

    @dday.command(name="ui", help="D-Day UI 버튼 표시")
    @commands.has_permissions(administrator=True)
    async def dday_ui(self, ctx_or_interaction):
        """D-Day 관리 UI 버튼 표시"""
        embed = discord.Embed(
            title="📅 D-Day 관리",
            description="아래 버튼을 클릭하여 D-Day를 관리하세요.",
            color=discord.Color.blue()
        )
        embed.add_field(
            name="기능",
            value="➕ **추가** - 새로운 D-Day 추가\n"
                  "📋 **목록** - 현재 D-Day 목록 보기\n"
                  "🔄 **새로고침** - D-Day 값 업데이트\n"
                  "❌ **삭제** - D-Day 삭제하기\n"
                  "📂 **채널 생성/동기화** - 카테고리/채널 만들기\n"
                  "🔁 **채널 새로고침** - 채널 이름 업데이트",
            inline=False
        )
        
        view = DDayView(self)
        
        if isinstance(ctx_or_interaction, discord.Interaction):
            await ctx_or_interaction.response.send_message(embed=embed, view=view, ephemeral=True)
        else:
            await ctx_or_interaction.send(embed=embed, view=view)

    @dday.command(name="추가", help="D-Day 추가")
    async def add_dday(self, ctx, name: str, date: str):
        """D-Day 추가"""
        try:
            target_date = datetime.strptime(date, "%Y-%m-%d")
            
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            # 새로운 행 추가
            row = ws.max_row + 1
            ws[f"A{row}"] = name
            ws[f"B{row}"] = target_date.date()
            ws[f"C{row}"] = self.calculate_dday(target_date)
            ws[f"D{row}"] = datetime.now().date()
            ws[f"E{row}"] = "활성"
            
            # 셀 스타일 설정
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            wb.save(self.excel_file)
            
            dday_count = self.calculate_dday(target_date)
            dday_text = f"D{dday_count}" if dday_count < 0 else f"D+{dday_count}"
            embed = discord.Embed(
                title="✅ D-Day 추가됨",
                description=f"**{name}**\n날짜: {date}\n{dday_text}",
                color=discord.Color.green()
            )
            await ctx.send(embed=embed)
        except ValueError:
            await ctx.send("❌ 날짜 형식이 잘못되었습니다. (YYYY-MM-DD)")
        except Exception as e:
            await ctx.send(f"❌ 오류: {e}")

    @dday.command(name="삭제", help="D-Day 삭제")
    async def delete_dday(self, ctx, *, name: str):
        """D-Day 삭제"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            found = False
            for row in range(2, ws.max_row + 1):
                if ws[f"A{row}"].value == name:
                    ws.delete_rows(row)
                    found = True
                    break
            
            if found:
                wb.save(self.excel_file)
                # 채널도 삭제
                if name in self.dday_channels:
                    try:
                        await self.dday_channels[name].delete()
                        del self.dday_channels[name]
                    except Exception as e:
                        print(f"채널 삭제 오류: {e}")
                await ctx.send(f"✅ **{name}** D-Day가 삭제되었습니다.")
            else:
                await ctx.send(f"❌ **{name}**을 찾을 수 없습니다.")
        except Exception as e:
            await ctx.send(f"❌ 오류: {e}")

    @dday.command(name="목록", help="D-Day 목록 표시")
    async def list_dday(self, ctx):
        """D-Day 목록 표시"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            if ws.max_row <= 1:
                await ctx.send("등록된 D-Day가 없습니다.")
                return
            
            embed = discord.Embed(title="📅 D-Day 목록", color=discord.Color.blue())
            
            for row in range(2, ws.max_row + 1):
                if not (name := ws[f"A{row}"].value):
                    continue

                date = ws[f"B{row}"].value or ""
                dday = ws[f"C{row}"].value
                
                if dday < 0:
                    text = f"📍 {dday}일 (경과)"
                elif dday == 0:
                    text = "🎉 D-Day!"
                else:
                    text = f"⏳ D+{dday}일"
                
                embed.add_field(
                    name=name,
                    value=f"{text}\n목표: {date}",
                    inline=False
                )
            
            await ctx.send(embed=embed)
        except Exception as e:
            await ctx.send(f"❌ 오류: {e}")

    @dday.command(name="공지", help="D-Day 공지 발송")
    @commands.has_permissions(manage_messages=True)
    async def announce_dday(self, ctx, channel: discord.TextChannel = None):
        """특정 채널에 D-Day 공지 발송"""
        if channel is None:
            settings = self.load_settings()
            channel_id = settings.get("dday_channel_id", 0)
            
            if channel_id == 0:
                await ctx.send("❌ D-Day 채널이 설정되지 않았습니다. `!채널설정 디데이 <채널>`로 설정하거나 `!디데이 공지 <채널>`로 지정해주세요.")
                return
            
            channel = self.bot.get_channel(channel_id)
            if not channel:
                await ctx.send("❌ 설정된 D-Day 채널을 찾을 수 없습니다.")
                return
        
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            embed = discord.Embed(
                title="📅 D-Day 공지",
                description="현재 진행 중인 D-Day 목록입니다.",
                color=discord.Color.gold()
            )
            
            for row in range(2, ws.max_row + 1):
                if not (name := ws[f"A{row}"].value):
                    continue

                dday = ws[f"C{row}"].value
                
                if dday < 0:
                    text = f"📍 {dday}일 (경과)"
                elif dday == 0:
                    text = "🎉 D-Day!"
                else:
                    text = f"⏳ D+{dday}일"
                
                embed.add_field(name=name, value=text, inline=True)
            
            await channel.send(embed=embed)
            await ctx.send(f"✅ D-Day 공지가 {channel.mention}에 발송되었습니다.")
        except Exception as e:
            await ctx.send(f"❌ 오류: {e}")

    @dday.command(name="새로고침", help="D-Day 업데이트")
    async def refresh_dday(self, ctx):
        """D-Day 값 업데이트"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            for row in range(2, ws.max_row + 1):
                if date_cell := ws[f"B{row}"].value:
                    if isinstance(date_cell, str):
                        date_cell = datetime.strptime(date_cell, "%Y-%m-%d")
                    ws[f"C{row}"] = self.calculate_dday(date_cell)
            
            wb.save(self.excel_file)
            await ctx.send("✅ D-Day가 새로고침되었습니다.")
        except Exception as e:
            await ctx.send(f"❌ 오류: {e}")

    def calculate_dday(self, target_date) -> int:
        """D-Day 계산"""
        if isinstance(target_date, str):
            target_date = datetime.strptime(target_date, "%Y-%m-%d")
        
        return (target_date - datetime.now()).days

    def load_settings(self):
        """설정 로드"""
        try:
            with open(self.settings_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {"news_channel_id": 0, "dday_channel_id": 0}

async def setup(bot):
    await bot.add_cog(DDayManager(bot))
