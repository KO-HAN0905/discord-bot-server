# -*- coding: utf-8 -*-
import discord
from discord.ext import commands
from discord.ui import View, Button, Modal, TextInput
from config import ADMIN_PASSWORD
import contextlib
import os
from datetime import datetime

# 관리자 비밀번호 인증 모달
class AdminPasswordModal(Modal, title="관리자 인증"):
    """관리자 비밀번호 인증"""
    
    password = TextInput(
        label="관리자 비밀번호",
        placeholder="비밀번호를 입력하세요",
        required=True,
        min_length=1,
        max_length=100
    )
    
    def __init__(self, callback):
        super().__init__()
        self.callback = callback
    
    async def on_submit(self, interaction: discord.Interaction):
        if self.password.value == ADMIN_PASSWORD:
            await self.callback(interaction, True)
        else:
            embed = discord.Embed(
                title="❌ 인증 실패",
                description="비밀번호가 올바르지 않습니다.",
                color=discord.Color.red()
            )
            await interaction.response.send_message(embed=embed, ephemeral=True)

# 서버 통계 UI 버튼 뷰
class ServerStatsView(View):
    """서버 통계 UI 버튼"""
    
    def __init__(self, cog):
        super().__init__(timeout=300)
        self.cog = cog
    
    @discord.ui.button(label="🔄 새로고침", style=discord.ButtonStyle.primary)
    async def refresh_button(self, interaction: discord.Interaction, button: Button):
        """통계 새로고침"""
        await self.cog.show_server_stats(interaction)
    
    @discord.ui.button(label="📊 상세정보", style=discord.ButtonStyle.secondary)
    async def detail_button(self, interaction: discord.Interaction, button: Button):
        """상세 정보 표시"""
        await self.cog.show_detailed_stats(interaction)
    
    @discord.ui.button(label="📅 D-Day", style=discord.ButtonStyle.success)
    async def dday_button(self, interaction: discord.Interaction, button: Button):
        """D-Day 정보 표시"""
        await self.cog.show_dday_list(interaction)


# 서버 통계 Cog
class ServerStats(commands.Cog):
    """서버 통계 기능"""
    
    def __init__(self, bot):
        self.bot = bot
        self.dday_file = "data/dday.xlsx"
    
    def get_dday_data(self):
        """D-Day 데이터 가져오기"""
        with contextlib.suppress(Exception):
            import openpyxl
            if not os.path.exists(self.dday_file):
                return []
            
            wb = openpyxl.load_workbook(self.dday_file)
            ws = wb.active
            dddays = []
            
            for row in range(2, ws.max_row + 1):
                name = ws[f"A{row}"].value
                target_date = ws[f"B{row}"].value
                message = ws[f"C{row}"].value
                
                if name and target_date:
                    with contextlib.suppress(Exception):
                        if isinstance(target_date, str):
                            target_date = datetime.strptime(target_date, "%Y-%m-%d")
                        
                        today = datetime.now()
                        d_day = (target_date.date() - today.date()).days
                        
                        dddays.append({
                            "name": name,
                            "date": target_date.strftime("%Y-%m-%d"),
                            "days": d_day,
                            "message": message or ""
                        })
            
            return sorted(dddays, key=lambda x: x["days"])
        return []
    
    def _count_channels(self, guild, channel_type):
        """채널 타입별 개수 반환"""
        return sum(isinstance(c, channel_type) for c in guild.channels)
    
    def _get_dday_status_text(self, dddays):
        """D-Day 상태 텍스트 생성"""
        dday_text = ""
        for dday in dddays[:3]:
            status = ("✅ 완료" if dday["days"] < 0 else 
                     "🔴 오늘" if dday["days"] == 0 else
                     f"🟡 {dday['days']}일 남음" if dday["days"] <= 7 else
                     f"🟢 {dday['days']}일 남음")
            dday_text += f"• **{dday['name']}**: {status}\n"
        
        if len(dddays) > 3:
            dday_text += f"• ... 외 {len(dddays) - 3}개"
        return dday_text
    
    async def _send_response(self, ctx_or_interaction, embed, view):
        """응답 전송"""
        is_interaction = isinstance(ctx_or_interaction, discord.Interaction)
        if is_interaction:
            await ctx_or_interaction.followup.send(embed=embed, view=view)
        else:
            await ctx_or_interaction.send(embed=embed, view=view)
    
    async def show_server_stats(self, ctx_or_interaction):
        """서버 통계 표시"""
        is_interaction = isinstance(ctx_or_interaction, discord.Interaction)
        guild = ctx_or_interaction.guild
        
        if is_interaction:
            await ctx_or_interaction.response.defer()
        
        if not guild:
            msg = "❌ 서버 정보를 가져올 수 없습니다."
            if is_interaction:
                await ctx_or_interaction.followup.send(msg)
            else:
                await ctx_or_interaction.send(msg)
            return
        
        # 기본 통계
        member_count = guild.member_count
        text_channels = self._count_channels(guild, discord.TextChannel)
        voice_channels = self._count_channels(guild, discord.VoiceChannel)
        role_count = len(guild.roles)
        boost_level = guild.premium_tier
        boost_count = guild.premium_subscription_count
        online_members = sum(m.status != discord.Status.offline for m in guild.members)
        
        # Embed 생성
        embed = discord.Embed(
            title=f"📊 {guild.name} 서버 통계",
            color=discord.Color.blue(),
            timestamp=datetime.now()
        )
        
        # 서버 아이콘
        if guild.icon:
            embed.set_thumbnail(url=guild.icon.url)
        
        # 기본 정보
        embed.add_field(
            name="👥 멤버 정보",
            value=f"전체 멤버: **{member_count}명**\n온라인: **{online_members}명**",
            inline=True
        )
        
        embed.add_field(
            name="📢 채널 정보",
            value=f"텍스트: **{text_channels}개**\n음성: **{voice_channels}개**",
            inline=True
        )
        
        embed.add_field(
            name="🎖️ 역할 정보",
            value=f"역할: **{role_count}개**",
            inline=True
        )
        
        embed.add_field(
            name="⭐ 부스트 정보",
            value=f"부스트 레벨: **Tier {boost_level}**\n부스트 수: **{boost_count}개**",
            inline=True
        )
        
        embed.add_field(
            name="📅 서버 생성일",
            value=f"<t:{int(guild.created_at.timestamp())}:F>",
            inline=True
        )
        
        embed.add_field(
            name="👑 서버 소유자",
            value=f"{guild.owner.mention}",
            inline=True
        )
        
        # D-Day 정보
        if (dddays := self.get_dday_data()):
            embed.add_field(
                name="🎯 D-Day 요약",
                value=self._get_dday_status_text(dddays),
                inline=False
            )
        
        # 버튼과 함께 응답
        view = ServerStatsView(self)
        await self._send_response(ctx_or_interaction, embed, view)
    
    async def show_detailed_stats(self, interaction: discord.Interaction):
        """상세 통계 표시"""
        if not (guild := interaction.guild):
            await interaction.response.send_message("❌ 서버 정보를 가져올 수 없습니다.", ephemeral=True)
        else:
            await interaction.response.defer()
            
            # 상세 정보 계산
            verified = guild.verification_level.name
            
            # 역할별 멤버 수
            top_roles = sorted(
                [(role.name, len(role.members)) for role in guild.roles if role.name != "@everyone"],
                key=lambda x: x[1],
                reverse=True
            )[:5]
            
            embed = discord.Embed(
                title=f"📈 {guild.name} 상세 통계",
                color=discord.Color.green(),
                timestamp=datetime.now()
            )
            
            if guild.icon:
                embed.set_thumbnail(url=guild.icon.url)
            
            # 서버 설정
            embed.add_field(
                name="🔐 보안 설정",
                value=f"인증 레벨: **{verified}**\n명시적 필터링: **{guild.explicit_content_filter.name}**",
                inline=False
            )
            
            # 인기 역할
            if top_roles:
                roles_text = "\n".join([f"• {name}: {count}명" for name, count in top_roles])
                embed.add_field(
                    name="🎖️ 상위 역할",
                    value=roles_text,
                    inline=False
                )
            
            # 기능 활성화
            if (features := getattr(guild, 'features', [])):
                features_text = ", ".join([f"✅ {feature}" for feature in features[:5]])
                if len(features) > 5:
                    features_text += f", ... 외 {len(features) - 5}개"
                embed.add_field(
                    name="⚙️ 활성화된 기능",
                    value=features_text,
                    inline=False
                )
            
            view = ServerStatsView(self)
            await interaction.followup.send(embed=embed, view=view)
    
    async def show_dday_list(self, interaction: discord.Interaction):
        """D-Day 목록 표시"""
        dddays = self.get_dday_data()
        
        embed = discord.Embed(
            title="📅 D-Day 목록",
            description=dddays or "등록된 D-Day가 없습니다.",
            color=discord.Color.purple() if dddays else discord.Color.orange(),
            timestamp=datetime.now() if dddays else None
        )
        
        for dday in dddays:
            days = dday["days"]
            
            status = ("✅ 완료됨 ({abs(days)}일 경과)" if days < 0 else
                     "🔴 오늘!" if days == 0 else
                     f"🟡 **{days}일** 남음" if days <= 7 else
                     f"🟢 **{days}일** 남음")
            emoji = ("✅" if days < 0 else
                    "🔴" if days == 0 else
                    "🟡" if days <= 7 else
                    "🟢")
            
            message = f"목표: {dday['date']}\n상태: {status}"
            if dday['message']:
                message += f"\n설명: {dday['message']}"
            
            embed.add_field(
                name=f"{emoji} {dday['name']}",
                value=message,
                inline=False
            )
        
        view = ServerStatsView(self)
        await interaction.response.send_message(embed=embed, view=view, ephemeral=False)
    
    @commands.command(name="서버정보", aliases=["stats", "서버통계", "statisticts"])
    @commands.has_permissions(administrator=True)
    async def server_stats(self, ctx):
        """서버 통계 표시 (!서버정보) - 관리자 전용"""
        embed = discord.Embed(
            title="🔐 관리자 인증 필요",
            description="아래 버튼을 클릭하여 비밀번호를 입력하세요.",
            color=discord.Color.gold()
        )
        view = StatsAuthView(self)
        await ctx.send(embed=embed, view=view)


# 통계 인증 버튼 뷰
class StatsAuthView(View):
    def __init__(self, cog):
        super().__init__()
        self.cog = cog
    
    @discord.ui.button(label="🔐 비밀번호 입력", style=discord.ButtonStyle.primary)
    async def auth_button(self, interaction: discord.Interaction, button: Button):
        async def show_stats(modal_interaction, authenticated):
            if authenticated:
                await self.cog.show_server_stats(modal_interaction)
        
        modal = AdminPasswordModal(show_stats)
        await interaction.response.send_modal(modal)


# Cog 등록
async def setup(bot):
    await bot.add_cog(ServerStats(bot))
